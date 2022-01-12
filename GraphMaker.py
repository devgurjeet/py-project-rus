import pandas as pd
import matplotlib.pyplot as plt
import re
import ListOfPatterns
from openpyxl.chart.text import RichText
from openpyxl.drawing.text import Paragraph, ParagraphProperties, CharacterProperties, Font
import openpyxl
import sys
import json
import datetime
import os
import time
from openpyxl import Workbook

remote_addresses = ['96.118.150.184', '96.118.25.182']

def get_filtered(the_file):
    final_data = []
    with open(the_file, 'r') as f:
        for line in f:
            #print ("Check 1: "+ str(line))
            #print(line.strip().encode('unicode_escape'))
            try:
                json_data = json.loads(line.strip().encode('unicode_escape').decode())
                time_local = datetime.datetime.strptime(json_data['time_local'],'%d/%b/%Y:%H:%M:%S +0000')
            # print(json_data['remote_addr'])
            # continue
                if json_data["remote_addr"] in remote_addresses and "/" != json_data["request_uri"] and "-" != json_data["request_uri"]:
                    temp_d = {}
                temp_d["request_method"] = json_data['request_method']
                temp_d["request_uri"] = json_data['request_uri']
                temp_d["time_local"] = time_local
                temp_d["request_time"] = json_data['request_time']
                temp_d["Match_count"] = 0
                final_data.append(temp_d)

            except Exception as ex:
                print(f"Error: {line} - {ex}")

            
    final_data.sort(key=lambda d: d["time_local"])
    return final_data


def check(datas, patterns, keyname):
    unmatched_data = []
    multiple_matched_data = []

    for data in datas:
        if data["Match_count"] == 0:
            unmatched_data.append(data["request_method"] + ":" + data["request_uri"])
        elif data["Match_count"] > 1:
            multiple_matched_data.append(data["request_method"] + ":" + data["request_uri"])
    print ("Pending APIs to match with patterns {}:".format(len(unmatched_data)))
    for a in unmatched_data:
        print (a)
    print ("APIs matching with multiple patterns:")
    for a in multiple_matched_data:
        print (a)

    if len(unmatched_data)==0 and len(multiple_matched_data)==0:
        return True
    else:
        return False

env_name = "DEV24"
app_name = "Simulator"
primary_file = os.getcwd() + "\Scenario1_1K_access.log"
#primary_file = os.getcwd() + "/log_files/scenario1.log.1"
env_excel_file = os.getcwd() + "/excel_files/" + env_name + ".xlsx"


if not os.path.isfile(primary_file):
    print("Source file path {} does not exist. Exiting...".format(primary_file))
    sys.exit()
else:
    print ("------------------------------------------------")
    print ("files are: ")
    print (primary_file)
    print (env_excel_file)
    print ("Environment is: " + env_name)
    print ("App is: " + app_name)
    print ("------------------------------------------------\n\n")


print ("------------------------------------------------")
print ("Getting date and API filtered data for primary...")
app_data_primary = get_filtered(primary_file)
print ("done {}".format(len(app_data_primary)))
#print ("done {}".format((app_data_primary)))
#import sys
#sys.exit()
#input()
print ("------------------------------------------------")

for pattern in ListOfPatterns.patterns:
    friendly_name = pattern["pattern"].replace(".*\\","<>").replace(".*", "<>").replace(".*\?", "<>").replace("$", "").replace("^", "").replace("\d*","<>").replace("(", "^(").replace("[\-A-Za-z0-9_]","<template_name>")
    #print (friendly_name)
    pattern["friendly_name"] = friendly_name
    pattern["regex"] = re.compile(pattern["pattern"], re.IGNORECASE)
print ("------------------------------------------------\n\n")

print ("------------------------------------------------")


print ("Matching patterns for primary...")
for data in app_data_primary:
    for pattern in ListOfPatterns.patterns:
         if re.match(pattern["regex"], data["request_method"] + ":" + data["request_uri"]) is not None:
             data["Match_count"] +=1
             data['Name'] = pattern['friendly_name']
             
             #print(f"data {pattern['friendly_name']} - {data}")
                
             pattern["api_details_primary"].append(data)
if check(app_data_primary, ListOfPatterns.patterns, "api_details_primary") == False:
    print("Some of the APIs could not find a match")
else:
    print ("All APIs found match")
print ("------------------------------------------------\n\n")


data = []
print ("------------------------------------------------")
print ("filling data....")
row_no = 1
chart_position = 2
for pattern in ListOfPatterns.patterns:
    #
    if len(pattern["api_details_primary"]) == 0:
        #print (".........does not have data for primary")
        continue
    else:
        print ("API:", pattern["friendly_name"])
        print (".........has data for primary={}".format(len(pattern["api_details_primary"])))
        print(pattern["api_details_primary"])
        data += pattern["api_details_primary"]
        


#pandas code.
df = pd.DataFrame(data)
df.request_time = df.request_time.astype('float')

mint = df.time_local.min()
maxt = df.time_local.max()

delta = datetime.timedelta(minutes=1)
difft = maxt - mint

start_date = mint
end_date = maxt

tempDf = pd.DataFrame()

all_frames = []
while start_date <= end_date:
    
    tempDf = pd.DataFrame()
    
    e_date = start_date + delta;
    mask = (df['time_local'] > start_date) & (df['time_local'] <= e_date) 
    #& (df.Name == 'GET:/vnms/dashboard/appliance/<>/live\?uuid=<>&command=interfaces/brief')
    
    print("==================================")
    print(f"{start_date} to {e_date}")
    tempDf = df[mask]
    #d = tempDF.groupby('Name').sum().reset_index()
    #print(tempDf)
#     tempDF = df[mask]
#     tempDF.groupby('Name').mean().reset_index()
    
    tempDf['start_date'] = start_date
    tempDf['end_date'] = e_date
    tempDf['range'] = f"{start_date} to {e_date}"
    
    all_frames.append(tempDf)
    
    print(len(tempDf))
    print("==================================")
    start_date += delta 

d = pd.concat(all_frames)

# grouping code.
ddf = d.groupby(['Name', 'start_date', 'end_date']).mean().reset_index()

if os.path.exists(env_excel_file):
    os.remove(env_excel_file)
env_workbook = Workbook()
app_sheet = env_workbook.create_sheet(app_name)
#chart_sheet = env_workbook.create_sheet("charts")

row_no = 1
chart_position = 2

for r in ddf.Name.unique():
    tempDf = ddf[ddf.Name == r]
    
        
    app_sheet.merge_cells('A{}:P{}'.format(row_no, row_no))
    app_sheet.cell(row=row_no, column=1).value = tempDf.iloc[0].Name
    row_no+=1
    
    app_sheet.cell(row=row_no, column=1).value = "Date"
    app_sheet.cell(row=row_no, column=2).value = "Primary"
    row_no+=1
    
    min_row = row_no
    for idx in  range(0, len(tempDf)):
        app_sheet.cell(row=row_no, column=1).value = f"{tempDf.iloc[idx].start_date} to {tempDf.iloc[idx].start_date}"
        app_sheet.cell(row=row_no, column=2).value = tempDf.iloc[idx].request_time
        
        row_no += 1
    
    row_no+=1
    
    chart = openpyxl.chart.LineChart()
    chart.title = tempDf.iloc[0].Name
    chart.style = 2
    chart.y_axis.title = 'Seconds'
    chart.x_axis.title = 'Date'
    font = Font(typeface='Verdana')
    size = 600 # 14 point size
    cp = CharacterProperties(latin=font, sz=size, b=False) # Not bold
    pp = ParagraphProperties(defRPr=cp)
    rtp = RichText(p=[Paragraph(pPr=pp, endParaRPr=cp)])
    chart.title.txPr = rtp

    data   = openpyxl.chart.Reference(app_sheet, min_col=2, min_row=min_row, max_row=row_no-2, max_col=2)
    labels = openpyxl.chart.Reference(app_sheet, min_col=1, min_row=min_row, max_row=row_no-2)
    
    #print(data)
    #print(labels)
    
    chart.add_data(data, titles_from_data=False)
    chart.set_categories(labels)
    chart.width = 30
    app_sheet.add_chart(chart, "C{}".format(chart_position))
    chart_position += 16

env_workbook.remove(env_workbook['Sheet'])
env_workbook.save(filename = 'env_file2.xlsx')
print("\n\n\n===================================================")
print("print: env_file2.xlsx")
print("===================================================\n\n\n")